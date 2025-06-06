from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def formatar_palavras_chave(file_path, palavras_chave, amarelo=False):
    wb = load_workbook(file_path)
    ws = wb.active

    fundo_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fundo_amarelo = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    letra_branca_negrito = Font(color="FFFFFF", bold=True)
    letra_preta_negrito = Font(color="000000", bold=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = fundo_preto
        cell.font = letra_branca_negrito

    for row in range(2, ws.max_row + 1):
        nome_vaga = ws.cell(row=row, column=5).value

        if nome_vaga:
            for palavra in palavras_chave:
                if palavra.lower() in nome_vaga.lower():

                    for col in range(1, 15):  # Colunas de A (1) até N (14)
                        cell = ws.cell(row=row, column=col)
                        if amarelo:
                            cell.fill = fundo_amarelo
                            letra_preta_negrito
                        else:
                            cell.fill = fundo_vermelho
                            cell.font = letra_branca_negrito
                    break

    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['F'].hidden = True
    wb.save(file_path)

    print("Planilha atualizada com as palavras-chave formatadas!")

